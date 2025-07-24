import os
import re
import streamlit as st
import pandas as pd
import pdfplumber
from docx import Document
import spacy
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Load NLP models
nlp = spacy.load("en_core_web_sm")
model = SentenceTransformer("all-MiniLM-L6-v2")

st.set_page_config(page_title="CV Screening Assistant", layout="centered")
st.title("🧠 Automated Hiring Assistant")

# ----------------------------
# Utility Functions
# ----------------------------

def extract_text_from_pdf(file):
    with pdfplumber.open(file) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)

def extract_text_from_docx(file):
    doc = Document(file)
    return "\n".join([p.text for p in doc.paragraphs])

def extract_text(file):
    if file.name.endswith(".pdf"):
        return extract_text_from_pdf(file)
    elif file.name.endswith(".docx"):
        return extract_text_from_docx(file)
    else:
        return file.read().decode("utf-8")

def extract_email(text):
    match = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', text)
    return match.group(0) if match else ""

def extract_phone(text):
    match = re.search(r'(\+?\d{1,3}[-.\s]?)?(\d{10}|\d{3}[-.\s]?\d{3}[-.\s]?\d{4})', text)
    return match.group(0) if match else ""

def remove_company_name(text, company):
    return text.replace(company, "[REDACTED]")

def get_similarity(text, jd_embedding):
    cv_embedding = model.encode(text, convert_to_tensor=True)
    return round(util.cos_sim(cv_embedding, jd_embedding).item() * 100, 2)

def extract_field(label_pattern, text):
    pattern = rf"(?i)(?:{label_pattern})\s*[:\-–—]\s*(.*)"
    try:
        match = re.search(pattern, text)
        return match.group(1).strip() if match else ""
    except Exception as e:
        print(f"Error extracting field '{label_pattern}': {e}")
        return ""

def extract_name(text):
    match = re.search(r"(?:Name)\s*[:\-–—]\s*([A-Z][a-z]+\s+[A-Z][a-z]+)", text)
    if match:
        return match.group(1).strip()
    lines = text.splitlines()
    for line in lines[:10]:
        doc = nlp(line)
        for ent in doc.ents:
            if ent.label_ == "PERSON":
                return ent.text.strip()
    return ""

def extract_role(text):
    return extract_field("Role|Position|Job Title|Designation", text)

def extract_company(text):
    return extract_field("Company|Currently Working at|Employer|Organisation|Working with|Currently At", text)

def generate_email(name, email, jd_filename):
    body = f"""\nSubject: Job Opportunity

Dear {name or 'Candidate'},

We are pleased to inform you that your profile has been shortlisted based on your resume. Please find the job description attached.

Best regards,
Recruitment Team
"""
    filename = f"{name or email}_email.txt"
    with open(filename, "w") as f:
        f.write(f"To: {email}\n\n{body}")
    return filename

# ----------------------------
# UI Inputs
# ----------------------------

client_name = st.text_input("🏢 Enter Client Name (Confidential)")

uploaded_jd = st.file_uploader("📌 Upload Job Description (TXT, PDF, DOCX)", type=["txt", "pdf", "docx"])
uploaded_resumes = st.file_uploader("📁 Upload Candidate Resumes", type=["pdf", "docx", "txt"], accept_multiple_files=True)

# ----------------------------
# Process Button
# ----------------------------

if st.button("🚀 Process All Resumes") and uploaded_jd and uploaded_resumes:
    jd_text = extract_text(uploaded_jd)
    redacted_jd = remove_company_name(jd_text, client_name)
    jd_embedding = model.encode(redacted_jd, convert_to_tensor=True)

    data = []

    for file in uploaded_resumes:
        text = extract_text(file)
        email = extract_email(text)
        phone = extract_phone(text)
        name = extract_name(text)
        role = extract_role(text)
        ctc = extract_field("CTC|Current CTC", text)
        ectc = extract_field("Expected CTC|ECTC", text)
        experience = extract_field("Experience|Total Experience", text)
        notice = extract_field("Notice|Notice Period", text)
        company = extract_company(text)

        score = get_similarity(text, jd_embedding)

        generate_email(name, email, uploaded_jd.name)

        resume_path = f"resume_{name or email}.txt"
        with open(resume_path, "w", encoding="utf-8") as f:
            f.write(text)

        data.append({
            "Client": client_name,
            "Name": name,
            "Email": email,
            "Phone": phone,
            "Role": role,
            "CTC": ctc,
            "ECTC": ectc,
            "Experience": experience,
            "Notice Period": notice,
            "Currently At": company,
            "Match %": score,
            "Status": "Profile Shared",
            "Resume File": resume_path
        })

    df = pd.DataFrame(data)

    st.subheader("📊 Candidate Matching Preview")
    st.dataframe(df.drop(columns=["Resume File"]))

 if st.button("💾 Generate Excel and Enable Resume Downloads"):
    filename = "Candidate_Tracker.xlsx"
    temp_df = df.drop(columns=["Resume File"])
    temp_df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    status_list = ["Profile Shared", "Shortlisted", "Interview L1", "Interview L2", "Offered", "Rejected"]
    dv = DataValidation(type="list", formula1=f'"{",".join(status_list)}"', allow_blank=True)
    ws.add_data_validation(dv)

    for row in range(2, len(temp_df) + 2):
        dv.add(ws[f"M{row}"])  # Assuming "Status" is column M

    wb.save(filename)

    with open(filename, "rb") as f:
        excel_bytes = f.read()
        st.download_button(
            label="⬇️ Download Candidate Tracker Excel",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

                )
